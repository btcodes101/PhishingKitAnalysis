import xlsxwriter
import os
import glob
import hashlib
from zipfile import ZipFile
import requests
from collections import namedtuple
import subprocess
import re
from datetime import datetime
import json
from pathlib import Path
import wget
from openpyxl import load_workbook
from openpyxl import Workbook
import html

WebsiteStatus = namedtuple('WebsiteStatus', ['status_code', 'reason'])

#function to get status of URL
def get_status(site):
    try:
        response = requests.head(site, timeout=5)
        status_code = response.status_code
        reason = response.reason
    except requests.exceptions.ConnectionError:
        status_code = '000'
        reason = 'ConnectionError'

    except :
        status_code = 'Unknown'
        reason = 'UnknownConnectionError'

    website_status = WebsiteStatus(status_code, reason)
    return website_status


#calculate SHA-1 of a zip file
def Sha1Hasher(file_path):

    buf_size = 65536
    sha1 = hashlib.sha1()

    with open(file_path, 'rb') as f:
        while True:
            data = f.read(buf_size)
            if not data:
                break
            sha1.update(data)

    return format(sha1.hexdigest())


list_of_all_kit_url_files=['/home/kali/phishfinder/kits.txt']#,'/home/kali/phishfinder_openphish/phishfinder/kits.txt','/home/kali/phishfinder_github/phishfinder/kits.txt','/home/kali/phishfinder-modified/kits.txt']
#list_of_all_kit_url_files.append('/home/kali/Phishing_Kits_phishunt/phishing_kits/Sep_2021/urls_new.txt')
#list_of_all_kit_url_files.append('/home/kali/phishfinder/kits/collected_kits/small_kits/ALL_KITS_OF_INTEREST/kits.txt')
list_of_all_kit_url_files.append('/home/kali/phishfinder/kits/collected_kits/small_kits/ALL_KITS_OF_INTEREST/openphish_kits.txt')
list_of_all_kit_url_files.append('/home/kali/phishfinder/kits/collected_kits/small_kits/ALL_KITS_OF_INTEREST/phishtank_kits.txt')
list_of_all_kit_url_files.append('/home/kali/phishfinder/kits/collected_kits/small_kits/ALL_KITS_OF_INTEREST/apwg_kits.txt')
list_of_all_kit_url_files.append('/home/kali/phishfinder/kits/collected_kits/small_kits/ALL_KITS_OF_INTEREST/urls_new.txt')
#list_of_all_kit_url_files.append('/home/kali/Phishing_Kits_phishunt/phishing_kits/Jul_2020/urls_new.txt')
#list_of_all_kit_url_files.append('/home/kali/Phishing_Kits_phishunt/phishing_kits/Jul_2021/urls_new.txt')
#list_of_all_kit_url_files.append('/home/kali/Phishing_Kits_phishunt/phishing_kits/Jun_2020/urls_new.txt')
#list_of_all_kit_url_files.append('/home/kali/Phishing_Kits_phishunt/phishing_kits/Jun_2021/urls_new.txt')
#list_of_all_kit_url_files.append('/home/kali/Phishing_Kits_phishunt/phishing_kits/Mar_2021/urls_new.txt')
#list_of_all_kit_url_files.append('/home/kali/Phishing_Kits_phishunt/phishing_kits/May_2020/urls_new.txt')
#list_of_all_kit_url_files.append('/home/kali/Phishing_Kits_phishunt/phishing_kits/May_2021/urls_new.txt')
#list_of_all_kit_url_files.append('/home/kali/Phishing_Kits_phishunt/phishing_kits/Nov_2020/urls_new.txt')
#list_of_all_kit_url_files.append('/home/kali/Phishing_Kits_phishunt/phishing_kits/Oct_2020/urls_new.txt')
#list_of_all_kit_url_files.append('/home/kali/Phishing_Kits_phishunt/phishing_kits/Sep_2020/urls_new.txt')
#list_of_all_kit_url_files.append('/home/kali/Phishing_Kits_Lab/openphish/kits.txt')
#list_of_all_kit_url_files.append('/home/kali/Phishing_Kits_Lab/phishtank/kits.txt')

#/home/kali/Phishing_Kits_phishunt/phishing_kits/Aug_2021/urls.txt"
KIT_DIRECTORY= "/home/kali/phishfinder/kits/collected_kits/small_kits/ALL_KITS_OF_INTEREST/"
PROGPILOT_OUTPUT=KIT_DIRECTORY+"PROGPILOT-OUTPUT/"
URL_ULTIMATE_LIST_FILE= KIT_DIRECTORY+"ultimate_url_list.txt"
MY_KIT_LIST="/home/kali/phishfinder/kits/collected_kits/small_kits/my_kit_list.txt"
PHP_MALWARE_FINDER = "/home/kali/phishfinder/kits/collected_kits/small_kits/php-malware-finder/php-malware-finder/phpmalwarefinder"
PDSCAN = "/home/kali/phishfinder/kits/collected_kits/small_kits/pdscan"
KIT_AUTHOR_LIST = "/home/kali/phishfinder/kits/collected_kits/small_kits/Phishing_Kit_Author_Indicators.txt"
PATH_TO_PHP = "/home/kali/php-7.3.24/sapi/cli/php"
PATH_TO_DEOBFUSCATOR= "/home/kali/phishfinder_automation/Sajjad-CLI-Script/Deobfuscator/PHPDeobfuscator/index.php"
PATH_TO_SAJJAD_SCRIPT="/home/kali/phishfinder_automation/Sajjad-CLI-Script/phish/main.php"
PATH_TO_MAIL_OUTOUT="/home/kali/phishfinder_automation/track_email/mail.out"

if not os.path.exists(PROGPILOT_OUTPUT):
    os.makedirs(PROGPILOT_OUTPUT)

f=open(URL_ULTIMATE_LIST_FILE,"w")

for x in list_of_all_kit_url_files:
	try:
		f1=open(x,"r")
		for y in f1:
			f.write(y.strip()+"\n")
		f1.close()
	except:
		pass

f.close()	

now = datetime.now()
date_time = now.strftime("%d.%m.%Y_%H.%M.%S")

workbook = Workbook()
sheet = workbook.active

sheet["A1"] = "SL No."
sheet["B1"] = "Name" 
sheet["C1"] = "SHA-1"
sheet["D1"] = "Valid Kit"
sheet["E1"] = "Source URL"
sheet["F1"] = "HTTP/HTTPS"
sheet["G1"] = "Reachability Status"
sheet["H1"] = "Kit Size"
sheet["I1"] = "File extensions inside kit"
sheet["J1"] = "Target Entity"
sheet["K1"] = "Storage in files"
sheet["L1"] = "File names"
sheet["M1"] = "Storage in DB"
sheet["N1"] = "Send Email Detected from dynamic analysis"
sheet["O1"] = "Email IDs in kit"
sheet["P1"] = "Email IDs from dynamic analysis"
sheet["Q1"] = "Obfuscation to hide emails detected"
sheet["R1"] = "Progpilot output"
sheet["S1"] = "Status"
sheet["T1"] = "Time of Review"
sheet["U1"] = "Send Email Detected from code"
sheet["V1"] = "Parameter names"
sheet["W1"] = "Analysed Live Data" 
sheet["X1"] = "Download URLs"
sheet["Y1"] = "Hardcoded Credentials"
sheet["Z1"] = "PHP Malware Scan"
sheet["AA1"] = "Telegram bot usage"
sheet["AB1"] = "Google Drive usage"
sheet["AC1"] = "Renamed kit"
sheet["AD1"] = "Hidden Emails"
sheet["AE1"] = "Infinite Loop"
sheet["AF1"] = "Probable Kit Authors"
sheet["AG1"] = "Is it a kit?"
sheet["AH1"] = "Telegram Bot Token, chat ID"
sheet["AI1"] = "Telegram Bot GetMe"
sheet["AJ1"] = "Telegram Bot GetChat"
sheet["AK1"] = "Hidden Telegram Bot Usage"
sheet["AL1"] = "Hidden Telegram Bots"

zip_files=[]

os.chdir(KIT_DIRECTORY)

MY_EXCEL_FILE_NAME = "PhishKitAnalysis-"+date_time+".xlsx"
#MY_EXCEL_FILE_NAME = "TelegramBotFinder-"+date_time+".xlsx"

workbook.save(filename=MY_EXCEL_FILE_NAME)

'''f = open(MY_KIT_LIST,"r")
for p in f:
	zip_files.append(p.strip())'''

for file1 in glob.glob("*.zip"):
    zip_files.append(file1)

for file1 in glob.glob("*.rar"):
    zip_files.append(file1)

for file1 in glob.glob("*.7z"):
    zip_files.append(file1)

for file1 in glob.glob("*.tar"):
    zip_files.append(file1)

index=2

zip_file_sizes = []

for y in zip_files:
	zip_file_sizes.append(os.stat(y).st_size)

sorted_zip_files = [x for _,x in sorted(zip(zip_file_sizes,zip_files))]

list_of_known_kit_authors = []
author_file = open(KIT_AUTHOR_LIST,"r")
for author in author_file:
	list_of_known_kit_authors.append(author.strip())
	
for kit in sorted_zip_files:
	print(kit)
	workbook = load_workbook(filename=MY_EXCEL_FILE_NAME)
	sheet = workbook.active	
	sheet["A"+str(index)] = str(index-1)	
	sheet["B"+str(index)] = kit		
	sheet["C"+str(index)] = str(Sha1Hasher(kit))

	is_valid_kit="TRUE"
	issue_with_name="FALSE"
	
	try:
		f = ZipFile(kit,"r")
	except:
		is_valid_kit="FALSE"
		print("issue with kit")	

	sheet["D"+str(index)] = str(is_valid_kit)
	
	original_name = kit
	if("(" in kit or "%" in kit or " " in kit or ")" in kit or "&" in kit or "$" in kit):
		new_name = kit.replace("(","_")
		new_name = new_name.replace(")","_")
		new_name = new_name.replace("%","_")
		new_name = new_name.replace("&","_")
		new_name = new_name.replace("$","_")		
		new_name = new_name.replace(" ","")
		kit = new_name
		os.rename(original_name,new_name)
		sheet["AC"+str(index)] = new_name
		
	if(is_valid_kit=="TRUE"): #we need to proceed with analysis only if zip is valid
		print("kit is valid")
		f=open(URL_ULTIMATE_LIST_FILE, "r")
		myurl=""
		for x in f:
			timestamp=x[:14]
			url=x[15:].strip()
			if(kit[:14]==timestamp and url.endswith(".zip")):# and original_name[15:] in url):
				myurl=url

		sheet["E"+str(index)]=str(myurl)
		if(myurl.startswith("https://")):
			sheet["F"+str(index)] = "HTTPS"					
		if(myurl.startswith("http://")):
			sheet["F"+str(index)] = "HTTP"
		if(len(myurl)>0):
			website_status = get_status(url)
			sheet["G"+str(index)]= str(website_status.status_code)+" "+ website_status.reason
		
		sheet["H"+str(index)]=str(os.stat(kit).st_size)
		
		#unzip the file
		s = "unar -d " + kit
		os.system(s)

		try:
			output = subprocess.check_output("cd "+kit[:-4].strip()+" && find . -type f | perl -ne 'print $1 if m/\.([^.\/]+)$/' | sort -u",shell=True)
		except:
			continue

		sheet["I"+str(index)] = output.decode("utf-8")
		
		if("html" in output.decode("utf-8") or "php" in output.decode("utf-8")):
			sheet["AG"+str(index)] = "TRUE"
		else:
			sheet["AG"+str(index)] = "FALSE"			

		PATH = KIT_DIRECTORY+kit[:-4]+"/"
		
		result = [os.path.join(dp, f) for dp, dn, filenames in os.walk(PATH) for f in filenames if os.path.splitext(f)[1] == '.php']

		list_of_telegram_api=[]
		list_of_google_drive=[]
		#change 17th May 2022 : added telegram bot part
		list_of_telegram_bots=[]
		
		for file1 in result:
			f = open(file1,"r")
			try:
				for line in f.readlines():			
					if("api.telegram.org" in line):
						if(line not in list_of_telegram_api):				
							list_of_telegram_api.append(line)
							print("Found a line containing telegram token.....",line) 
							#worksheet.write('K'+str(index),"TRUE")
							re_line = re.search('bot\d+:AA[a-zA-Z0-9-_]{20,60}', line)
							if("chat_id" in line):
								re_line1 = re.search('(bot.+?)/sendMessage\?chat_id=(.+?)&', line)
								list_of_telegram_bots.append(re_line1.group(1)+"$"+re_line1.group(2))
							else:	
								list_of_telegram_bots.append(re_line.group(0))
								
					if("drive.google.com" in line):
					#if("fopen" in line or "file_put_contents" in line and file1 not in php_files_of_interest):
						if(line not in list_of_google_drive):
							list_of_google_drive.append(line) 
						#worksheet.write('K'+str(index),"TRUE")
								
					#if("fopen" in line or "file_put_contents" in line and file1 not in php_files_of_interest):
					
			except Exception as e:
				print(str(e))

		sheet["AA"+str(index)] =  str(list_of_telegram_api)
		sheet["AB"+str(index)] =  str(list_of_google_drive)

#sheet["AH1"] = "Telegram Bot Token, chat ID"
#sheet["AI1"] = "Telegram Bot GetMe"
#sheet["AJ1"] = "Telegram Bot GetChat"
#sheet["AK1"] = "Hidden Telegram Bot Usage"
#sheet["AL1"] = "Hidden Telegram Bots"
		
		sheet["AH"+str(index)] = str(list_of_telegram_bots)
		#for bot1 in list_of_telegram_bots:
		#	if("$" in bot):
		#		sheet["AH"+str(index)]		
			
		#	else
				
		detected_kit_authors = []	

		for file1 in result:
			f = open(file1,"r")
			try:
				for line in f.readlines():			
					for auth in list_of_known_kit_authors:
						if auth in line:
							if auth not in detected_kit_authors:
								detected_kit_authors.append(auth)				
					
			except Exception as e:
				print(str(e))

		sheet["AF"+str(index)] = ','.join(sorted(detected_kit_authors))



		sheet["K"+str(index)]= "FALSE"

		names_of_info_files=[]

		php_files_of_interest=[]


		for file1 in result:
			f = open(file1,"r")
			try:
				for line in f.readlines():			
					if(file1 not in php_files_of_interest):
					#if("fopen" in line or "file_put_contents" in line and file1 not in php_files_of_interest):
						php_files_of_interest.append(file1) 
						#worksheet.write('K'+str(index),"TRUE")		
					
			except Exception as e:
				print(str(e))


		sheet["U"+str(index)] = "FALSE"

		for file1 in result:
			f = open(file1,"r")
			try:
				for line in f.readlines():
					line = line.replace(" ","")
					if ("mail(" in line):
						sheet["U"+str(index)] = "TRUE"
			except Exception as e:
				print(str(e))
		
		#print(result)
		set_of_emails_in_code=[]
		for file1 in result:
			f = open(file1,"r")
			try:
				for line in f.readlines():
					emails = re.findall("([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)", line)
					#print(emails)
					if(len(emails)>0):
						for x in emails:
							if(x not in set_of_emails_in_code):
								set_of_emails_in_code.append(x)
								#print("founf",x)	


			except Exception as e:
				print(str(e))
		
		emails_in_code_in_string  = ','.join(sorted(set_of_emails_in_code))
		sheet["O"+str(index)] = emails_in_code_in_string


		#print(php_files_of_interest)
		list_of_files_created=[]
		set_of_emails=[]
		list_of_variables_of_interest=[]
		#make the kit related files unremovable, to negate the effect of any file deletion script
		#os.system("chmod -r "+kit[:-4])
		#os.system("chmod -r "+kit)
		list_of_telegram_api_in_deobfuscated_code=[]
		for file1 in php_files_of_interest:
			SRC_PHP_FILE_PATH=file1.strip()
			DEOBFUSCATED_FILE_PATH = SRC_PHP_FILE_PATH					 		
			#print(file1)
			#deobfuscate the file syntax -> php deobfuscate/index.php -f src.php  > output.php
			fi = open(PATH_TO_MAIL_OUTOUT, "w") 
			fi.write("")#make the mail output file empty.. later on we will check if the php file triggers a change in mail output file
			fi.close()

			if("delete.php" in file1):
				continue

			if("guzzlehttp" in file1):
				continue

			if("/react/promise" in file1):
				continue		


			if("/vendor/swiftmailer" in file1):

				continue		
			try:
				y = os.popen(PATH_TO_PHP+" "+PATH_TO_DEOBFUSCATOR+" "+"-f "+SRC_PHP_FILE_PATH).read()
				#cmd=[PATH_TO_PHP, PATH_TO_DEOBFUSCATOR, "-f", SRC_PHP_FILE_PATH]
				#process = subprocess.Popen(cmd,stdout=subprocess.PIPE,stderr=subprocess.STDOUT)
				#y, errs = process.communicate(timeout=15).decode('utf-8')
			except:
				continue
			#print("deobfuscated "+file1)
			if("Fatal error" not in y):
				#print(y)
				try:
					f = open(SRC_PHP_FILE_PATH,"w")
					f.write(y)
					f.close()
				
				except Exception as e:
					print(str(e)) 				

			####hidden telegram bot check######			
			sf = open(SRC_PHP_FILE_PATH,"r")

			try:
				for line in sf.readlines():			
					if("api.telegram.org" in line):
						if(line not in list_of_telegram_api_in_deobfuscated_code):
							list_of_telegram_api_in_deobfuscated_code.append(line)
						
			except Exception as e:
				print(str(e))	
			####hidden telegram bot check end######
											
			DEOBFUSCATED_MODIFIED_FILE_PATH = DEOBFUSCATED_FILE_PATH+"-saj.php"
			w = PATH_TO_PHP+" "+PATH_TO_SAJJAD_SCRIPT+" "+DEOBFUSCATED_FILE_PATH+" "+DEOBFUSCATED_MODIFIED_FILE_PATH
			#cmd1 = [PATH_TO_PHP, PATH_TO_SAJJAD_SCRIPT, DEOBFUSCATED_FILE_PATH, DEOBFUSCATED_MODIFIED_FILE_PATH]
			#print(w)
			try:
				x = os.popen(w).read().strip()
				#process = subprocess.Popen(cmd1,stdout=subprocess.PIPE,stderr=subprocess.STDOUT)
				#x, errs = process.communicate(timeout=15).decode('utf-8')

			except:
				continue
			if(x.startswith("[") is False): #skip if sajjad script did not run propery
				continue

			#avoid a infinite loop situation which occurs for some phishing kits which try to create some random folders (code seems to be broken for such kits)
			READDIR_DETECTED = "FALSE"
			COPY_DETECTED = "FALSE"

			ddd = ""
			try:
				ddd = open(DEOBFUSCATED_FILE_PATH+"-saj.php",encoding="ISO-8859-1")
			except:
				continue
			for opo in ddd:
				if("readdir(" in opo):
					READDIR_DETECTED = "TRUE"
				
				if("copy(" in opo):
					COPY_DETECTED = "TRUE"

			
			if(READDIR_DETECTED == "TRUE" and COPY_DETECTED == "TRUE"):
				sheet["AE"+str(index)] = "TRUE"
				continue
	
			#iterate over the variables it has and create the arguments

			list_of_variables=[]
			for v in json.loads(x):
				list_of_variables.append(v)
			
			string_of_parameters="" #here we have set each parameter as hello.. we can also set static value for each parameter type like email name, cvv, credit card
			for x in list_of_variables:
				string_of_parameters = string_of_parameters + x + "="+ "hello "			
			#print(string_of_parameters)

			u = string_of_parameters+PATH_TO_PHP+" "+DEOBFUSCATED_MODIFIED_FILE_PATH
			
			print(string_of_parameters+PATH_TO_PHP+" "+DEOBFUSCATED_MODIFIED_FILE_PATH)


			try:
				#cmd2 = [string_of_parameters,"php",DEOBFUSCATED_MODIFIED_FILE_PATH]
				#process = subprocess.Popen(cmd2,stdout=subprocess.PIPE,stderr=subprocess.STDOUT)
				#eachfileoutput, errs = process.communicate(timeout=15)
				eachfileoutput = os.system(string_of_parameters+"php"+" "+DEOBFUSCATED_MODIFIED_FILE_PATH).read()

			#add check for command injection

			except Exception as e:
				print(str(e))
			#after executing each php file, check if mail.out file has been edited

			fi = open(PATH_TO_MAIL_OUTOUT,"r")
			
			if(os.stat(PATH_TO_MAIL_OUTOUT).st_size > 4):
				#os.system("cp /home/kali/phishfinder_automation/track_email/mail.out /home/kali/phishfinder_automation/track_email/sail.out")
				#print("YAAAY")
				#find all email ids.. mark send mail to true for this kit 
				for line in fi.readlines():
					if(line.startswith("To:")):
						emails = re.findall("([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)", line)
						if(len(emails)>0):
							for x in emails:
								if(x not in set_of_emails):
									set_of_emails.append(x)
			
	 				
			#print("here is the file: ",file1)
			#print("these are the email recipients :",set_of_emails)
			if(os.path.exists(KIT_DIRECTORY+"F1ls.txt")):
				for pp in list_of_variables:
					if pp not in list_of_variables_of_interest:
						list_of_variables_of_interest.append(pp)
				#print("confirmed that the kit "+kit+" stores info in files")
				sheet["K"+str(index)]= "TRUE"

				f=open(KIT_DIRECTORY+"F1ls.txt","r")
				
				for w in f:
					p=os.sep
					mypath=p.join(SRC_PHP_FILE_PATH.split(p)[:-1])+p	
					if(mypath+w.strip() not in list_of_files_created):
						list_of_files_created.append(mypath+w.strip())
						
					#print(w)
					
					#r = open(SRC_PHP_FILE_PATH.rsplit('/', 1)[0]+"/"+w.strip(),"r")
					#r = open(KIT_DIRECTORY+w.strip(),"r")
					#for d in r:
						#print(d)
					#temp_path= SRC_PHP_FILE_PATH.rsplit('/', 1)[0]+"/"+w.strip()
					temp_path = KIT_DIRECTORY+w.strip()
					if(os.path.exists(temp_path) and "F1ls.txt" not in temp_path):
						try:						
							os.remove(temp_path)
						except Exception as e:
							print(str(e))		
				try:
					os.remove(KIT_DIRECTORY+"F1ls.txt")
				except:
					print("")					
		list_of_files_created_string = ','.join(list_of_files_created)
		sheet["L"+str(index)]=list_of_files_created_string

		#os.system("chmod +r "+kit[:-4])
		#os.system("chmod +r "+kit)

		#try to fetch the live URL, run pdscan and store output of pdscan in excel
		p=os.sep
		base_url = p.join(myurl.split(p)[:-1])+p
		

		list_of_possible_download_urls = []

		#print("my base url is ->??",base_url)
		any_file_present="FALSE"
		for each_file in list_of_files_created:
			l = each_file.split(p)
			ind = l.index(kit[:-4])
			possible_url = base_url + p.join(l[ind+1:])
			print("possible_url : ",possible_url)	
			try:
				headers=requests.head(possible_url).headers
				print(headers)
				if('text' in headers.get('Content-Type', '')):
					now = datetime.now()
					if(possible_url not in list_of_possible_download_urls):
						list_of_possible_download_urls.append(possible_url)
					just_date = now.strftime("%d.%m.%Y")
					any_file_present = "TRUE"
					path_to_store_data = "/home/kali/phishfinder_automation/live_info/"+kit[:-4]+"/"+just_date
					Path(path_to_store_data).mkdir(parents=True, exist_ok=True)
					#wget.download(possible_url,out = path_to_store_data)
					possible_url= possible_url.replace("/var/html", "") #some kits store files with respect to /var/html
					my_exposed_file_dataset = open("urls_of_exposed_files_final.txt","a")
					my_exposed_file_dataset.write(possible_url.strip()+"\n")
					my_exposed_file_dataset.close()
					os.system("wget -P "+path_to_store_data+" "+possible_url)
			
			except Exception as e:
				print(str(e))
		
		if(any_file_present=="TRUE"):
			t = os.popen(PDSCAN + " file://"+ "/home/kali/phishfinder_automation/live_info/"+kit[:-4]+"/ --show-all").read()
			sheet["W"+str(index)] = t
			os.system("shred -uvz -n 6 /home/kali/phishfinder_automation/live_info/"+kit[:-4]+"/*")
			sheet["X"+str(index)] = ','.join(list_of_possible_download_urls)		

		list_of_variables_of_interest_string = ','.join(list_of_variables_of_interest)
		sheet["V"+str(index)] = list_of_variables_of_interest_string
			
		#pass the arguments to the new php file, and check if Fl1.txt file has name of any file created 

		#verify if that any file has indeed been created

		#record the file name in excel sheet		
		
		sheet["M"+str(index)] = "FALSE"

		result = [os.path.join(dp, f) for dp, dn, filenames in os.walk(PATH) for f in filenames if os.path.splitext(f)[1] == '.php']
		result_1 = [os.path.join(dp, f) for dp, dn, filenames in os.walk(PATH) for f in filenames if os.path.splitext(f)[1] == '.html']	
		result_2 = result+result_1

		for file1 in result:
			f = open(file1,"r")
			try:
				for line in f.readlines():					
					if ("mysql" in line or "SQLite3" in line):
						sheet["M"+str(index)] = "TRUE"
					
			except Exception as e:
				print(str(e))
		
		list_of_titles=[]
		for file1 in result_2:
			f = open(file1,"r")
			try:
				for line in f.readlines():
					lower_line = html.unescape(line.lower())					
					if ("<title>" in lower_line):
						matches = re.findall(r'<title>(.+?)</title>',lower_line)
						for pq in matches:
							if pq not in list_of_titles:
								list_of_titles.append(pq)
					
			except Exception as e:
				print(str(e))

		'''for file1 in result_1:
			f = open(file1,"r")
			try:
				for line in f.readlines():
					lower_line = html.unescape(line.lower())				
					if ("<title>" in lower_line):
						matches = re.findall(r'<title>(.+?)</title>',lower_line)
						for pq in matches:
							if pq not in list_of_titles:
								list_of_titles.append(pq)
						
			except Exception as e:
				print(str(e))
		'''
		
		sheet["J"+str(index)] = ','.join(sorted(list(set(list_of_titles))))

		sheet["N"+str(index)] =  "FALSE"
		
		emails_in_string = ','.join(sorted(set_of_emails))
		#print(emails_in_string)
		if("@" in emails_in_string):		
			sheet["N"+str(index)] = "TRUE"						
			sheet["P"+str(index)] = emails_in_string

		##hidden telegram bot start##
		print("##########normal -> ",list_of_telegram_api_in_deobfuscated_code)
		print("##########deobfuscated -> ",list_of_telegram_api_in_deobfuscated_code)

		hidden_bot_lines = list(set(list_of_telegram_api_in_deobfuscated_code)-set(list_of_telegram_api))
		if(len(hidden_bot_lines)==0):
			print("No obfuscated code is being used to send telegram messages")
			sheet["AK"+str(index)] = "FALSE"
		else:
			sheet["AK"+str(index)] = "TRUE"
			print("Obfuscated code is being used to send telegram messages")
			sheet["AL"+str(index)] =  ','.join(sorted(hidden_bot_lines))

		##hidden telegram bot end##	

		hidden_emails = list(set(set_of_emails)-set(set_of_emails_in_code))
		if(len(hidden_emails)==0):
			print("No obfuscated code is being used to transmit emails")
			sheet["Q"+str(index)] = "FALSE"
		else:
			print("Obfuscated code is being used to transmit emails")
			sheet["Q"+str(index)] = "TRUE"
			sheet["AD"+str(index)] =  ','.join(sorted(hidden_emails))
		
		print("[*****************************CHECKPOINT*********************************]Run progpilot")

		kit_name=kit[:-4]

		if(os.path.exists(PROGPILOT_OUTPUT+kit[:-4].strip()+"_progpilot.txt")==False):
			os.system("progpilot "+kit[:-4]+">"+PROGPILOT_OUTPUT+kit[:-4].strip()+"_progpilot.txt")

		#vulnerabilities detected in progpilot

		f = open(PROGPILOT_OUTPUT+kit[:-4].strip()+"_progpilot.txt")
		list_vulns=[]
		for x in f:
			if("vuln_name" in x):
				y = x.split(":")[1].strip()
				if(y[:-1] not in list_vulns):
					list_vulns.append(y[:-1])

		list_vulns_in_string = ','.join(sorted(list_vulns))
		sheet["R"+str(index)] = list_vulns_in_string


		#find hardcoded credentials and secrets in kit using whispers

		hidden_creds = os.popen("whispers "+ kit[:-4]+"/").read()
		sheet["Y"+str(index)] = hidden_creds


		#scan kit for php webshells

		phpmalwarescan_output = os.popen(PHP_MALWARE_FINDER+" "+kit[:-4]).read()			
		sheet["Z"+str(index)] = phpmalwarescan_output
				
		sheet["S"+str(index)] = "DONE"
		now = datetime.now()
		date_time = now.strftime("%d/%m/%Y, %H:%M:%S")
		sheet["T"+str(index)] = date_time
			
		#remove kit folder after work is complete

		if(os.path.exists(KIT_DIRECTORY+kit[:-4])):
			p="rm -r "+KIT_DIRECTORY+kit[:-4]
			os.system(p)
		if(os.path.exists(KIT_DIRECTORY+kit)):
			p="rm "+KIT_DIRECTORY+kit
			os.system(p)

	else:
		if(issue_with_name=="TRUE"):
			sheet["S"+str(index)] = "ISSUE WITH KIT NAME"
		else:
			sheet["S"+str(index)] = "FAULTY KIT"

		now = datetime.now()
		date_time = now.strftime("%d/%m/%Y, %H:%M:%S")
		sheet["T"+str(index)] = date_time
		
	print("Kit Number :",str(index-1))
	index=index+1
	workbook.save(filename=MY_EXCEL_FILE_NAME)
workbook.close()
